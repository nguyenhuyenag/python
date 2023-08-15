from datetime import datetime
import openpyxl

def to_mysql_datetime(original_date):
    original_date = datetime.strptime(original_date, '%d/%m/%Y')
    return original_date.strftime('%Y-%m-%d')

# Load the Excel workbook
workbook = openpyxl.load_workbook(r'TCT_DSCTY_Ruirovehoadon_2023_DAKT.xlsx')

sheet = workbook['Sheet']  # Replace 'Sheet1' with your sheet name

column_change = 5   # Start from 0

# 1 is header
for (index, row) in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    if row[column_change]:  # Check if the cell is not empty
        # Create new value
        new_value = to_mysql_datetime(row[column_change])
        # Update
        sheet.cell(row=index, column=column_change+1, value=new_value)

# Save the changes to the Excel file
workbook.save(r'TCT_DSCTY_Ruirovehoadon_2023_DAKT.xlsx')

# Close the workbook
workbook.close()
