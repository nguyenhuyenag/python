import openpyxl
from datetime import datetime


def to_mysql_datetime(str_date):
    str_date = datetime.strptime(str_date, '%d/%m/%Y')
    return str_date.strftime('%Y-%m-%d')


# Load the Excel workbook
workbook = openpyxl.load_workbook(r'TCT_DSCT.xlsx')
sheet_name = workbook['Sheet']  # Replace 'Sheet1' with your sheet name

column_index = 5  # Start from 0

# Because 1 is Header so start = 2
for (index, row) in enumerate(sheet_name.iter_rows(min_row=2, values_only=True), start=2):
    if row[column_index]:
        # Create new value
        new_value = to_mysql_datetime(row[column_index])
        # Update
        sheet_name.cell(row=index, column=column_index + 1, value=new_value)

# Save the changes to the Excel file
workbook.save(r'TCT_DSCT.xlsx')

# Close the workbook
workbook.close()
