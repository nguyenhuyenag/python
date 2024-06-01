import openpyxl
import subprocess


def open_excel_with_office(file_path):
    try:
        subprocess.Popen(["start", "excel", file_path], shell=True)
    except Exception as ex:
        print(f"Error: {ex}")


# Load the Excel workbook
file_path = r'D:\WORK\Cloud\FirstDrive\Zother\chi-tieu.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet_name = workbook['Th5']  # Replace 'Sheet1' with your sheet name

b_index = 1  # 0 -> A, 1 -> B,...
c_index = 2

try:
    # Vì 0 is Header (A, B, C,...) nên start = 1
    for row, column in enumerate(sheet_name.iter_rows(min_row=0, values_only=True), start=1):
        column_value = str(column[b_index])
        # print(column_value)
        # Check if the column_value is not None before using eval
        if column_value is not None:
            try:
                evaluated_value = int(column_value) if column_value.isdigit() else eval(column_value)
                sheet_name.cell(row=row, column=c_index + 1, value=evaluated_value)
            except Exception as e:
                print(f"Error evaluating cell at row {row}, column {b_index}: {e}")

finally:
    workbook.save(file_path)
    workbook.close()
    open_excel_with_office(file_path)
